#!/usr/bin/env python3
import argparse
import csv
import math
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple


IDENT_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_$]*$")


DEFAULT_SYSTEMVERILOG_TEMPLATE = """\
`default_nettype none

module {{module_name}} (
{{port_declarations}}
);

{{state_declarations}}

{{timer_declarations}}

{{state_register_block}}

{{timer_block}}

{{comb_block}}

{{assertion_block}}

endmodule

`default_nettype wire
"""


DEFAULT_VERILOG_TEMPLATE = """\
module {{module_name}} (
{{port_declarations}}
);

{{state_localparams}}

reg [STATE_W-1:0] {{state_reg_name}};
reg [STATE_W-1:0] {{next_state_name}};

{{timer_declarations}}

{{state_register_block}}

{{timer_block}}

{{comb_block}}

endmodule
"""


# Kept as a public alias for callers that imported the old constant.
DEFAULT_TEMPLATE = DEFAULT_SYSTEMVERILOG_TEMPLATE


@dataclass
class Field:
    name: str
    width: str = "1"
    default: str = "0"

    @property
    def decl_width(self) -> str:
        width = str(self.width).strip()
        if width in ("", "1", "1'b1"):
            return ""
        if width.startswith("[") and width.endswith("]"):
            return f" {width}"
        if width.isdigit():
            value = int(width)
            if value <= 1:
                return ""
            return f" [{value - 1}:0]"
        return f" {width}"


@dataclass
class State:
    raw_name: str
    ident: str
    code: Optional[str] = None
    timer_limit: Optional[str] = None


@dataclass
class Transition:
    state: str
    condition: str
    next_state: str
    outputs: Dict[str, str] = field(default_factory=dict)
    row_num: int = 0

    @property
    def is_default(self) -> bool:
        return self.condition.strip().lower() in ("", "default", "else", "1", "1'b1", "true")


@dataclass
class Spec:
    meta: Dict[str, str]
    states: List[State]
    inputs: List[Field]
    outputs: List[Field]
    transitions: List[Transition]
    template: str = DEFAULT_TEMPLATE


def die(message: str) -> None:
    print(f"ERROR: {message}", file=sys.stderr)
    raise SystemExit(1)


def warn(message: str) -> None:
    print(f"WARNING: {message}", file=sys.stderr)


def read_csv_rows(path: Path) -> List[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as fp:
        reader = csv.DictReader(fp)
        if not reader.fieldnames:
            return []
        rows = []
        for row in reader:
            normalized = {}
            for key, value in row.items():
                if key is None:
                    continue
                normalized[key.strip().lower()] = "" if value is None else str(value).strip()
            if any(value != "" for value in normalized.values()):
                rows.append(normalized)
        return rows


def read_template_csv(path: Path) -> str:
    lines = []
    with path.open(newline="", encoding="utf-8-sig") as fp:
        reader = csv.reader(fp)
        for row in reader:
            if not row:
                lines.append("")
            else:
                lines.append(row[0])
    return "\n".join(lines)


def read_xlsx_sheets(path: Path) -> Tuple[Dict[str, List[Dict[str, str]]], Optional[str]]:
    try:
        import openpyxl
    except ImportError:
        die("Reading .xlsx needs openpyxl. Install with: python3 -m pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets: Dict[str, List[Dict[str, str]]] = {}
    template = None

    for ws in wb.worksheets:
        sheet_name = ws.title.strip().lower()
        values = list(ws.iter_rows(values_only=True))
        if sheet_name == "template":
            lines = []
            for row in values:
                first = "" if not row or row[0] is None else str(row[0])
                lines.append(first)
            candidate = "\n".join(lines).rstrip() + "\n"
            if "{{port_declarations}}" in candidate or "{{comb_block}}" in candidate:
                template = candidate
            continue

        non_empty = [row for row in values if row and any(cell is not None and str(cell).strip() for cell in row)]
        if not non_empty:
            sheets[sheet_name] = []
            continue

        headers = ["" if cell is None else str(cell).strip().lower() for cell in non_empty[0]]
        rows = []
        for raw_row in non_empty[1:]:
            row = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                cell = raw_row[idx] if idx < len(raw_row) else ""
                row[header] = "" if cell is None else str(cell).strip()
            if any(value != "" for value in row.values()):
                rows.append(row)
        sheets[sheet_name] = rows

    return sheets, template


def create_fsm_template(path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from openpyxl.workbook.defined_name import DefinedName
    except ImportError:
        die("Creating .xlsx templates needs openpyxl. Install with: python3 -m pip install openpyxl")

    wb = openpyxl.Workbook()
    guide = wb.active
    guide.title = "Guide"
    config = wb.create_sheet("Config")
    inputs = wb.create_sheet("Inputs")
    outputs = wb.create_sheet("Outputs")
    states = wb.create_sheet("States")
    transitions = wb.create_sheet("Transitions")

    guide_rows = [
        ["SYSTEMVERILOG FSM GENERATOR", "Workbook source for one generated FSM"],
        ["Quick workflow", "1. Edit Config  2. Add Inputs/Outputs  3. Add States  4. Add Transitions  5. Run make gen"],
        ["Important", "Yellow cells are user-editable. Do not rename sheets or header columns."],
        ["Transition priority", "Rows are evaluated from top to bottom. Put the highest-priority condition first and default last."],
        ["State timing", "Set timer_limit in States, then use timer_done in a transition condition."],
        ["Bus width", "Enter width as a number: 1, 4, 8, 16, 32. Width 1 generates a scalar."],
        ["Outputs", "Defaults come from Outputs. Branch overrides use name=value separated by semicolons."],
        ["Generate", f"python3 fsm_gen.py {path.name} -o generated"],
        ["Documentation", "See docs/USER_GUIDE.md for complete rules and troubleshooting."],
    ]
    for row in guide_rows:
        guide.append(row)

    config.append(["key", "value", "description"])
    config_rows = [
        ["module_name", "timed_controller_fsm", "Generated RTL module name"],
        ["language", "systemverilog", "systemverilog (recommended) or verilog"],
        ["clock", "clk", "Clock input name"],
        ["reset", "rst_n", "Reset input name"],
        ["reset_active_low", "true", "true: active-low; false: active-high"],
        ["reset_type", "async", "async or sync reset implementation"],
        ["reset_state", "IDLE", "State selected after reset"],
        ["encoding", "binary", "binary, onehot, or gray automatic state encoding"],
        ["state_reg_name", "state", "Internal current-state register"],
        ["next_state_name", "next_state", "Internal next-state signal"],
        ["timer_width", "8", "Shared state-timer width"],
        ["clock_enable", "", "Optional FSM/timer clock-enable input name"],
        ["assertions", "true", "Emit simulation-only SystemVerilog assertions"],
    ]
    for row in config_rows:
        config.append(row)

    inputs.append(["name", "width", "description"])
    for row in [
        ["start", "1", "Start the sequence"],
        ["abort", "1", "Return to error/idle path"],
        ["delay_cycles", "8", "Programmable WAIT duration"],
    ]:
        inputs.append(row)

    outputs.append(["name", "width", "default", "description"])
    for row in [
        ["busy", "1", "1'b0", "FSM is processing"],
        ["done", "1", "1'b0", "One-cycle completion indication"],
        ["error", "1", "1'b0", "Abort/error indication"],
    ]:
        outputs.append(row)

    states.append(["state", "code", "timer_limit", "description"])
    for row in [
        ["IDLE", "", "", "Wait for start"],
        ["WAIT", "", "delay_cycles", "Timed processing state"],
        ["DONE", "", "", "Report completion"],
        ["ERROR", "", "", "Report abort/error"],
    ]:
        states.append(row)

    transitions.append(["state", "condition", "next_state", "outputs", "description"])
    for row in [
        ["IDLE", "start", "WAIT", "busy=1'b1", "Start"],
        ["IDLE", "default", "IDLE", "", "Stay idle"],
        ["WAIT", "abort", "ERROR", "error=1'b1", "Abort has highest priority"],
        ["WAIT", "timer_done", "DONE", "busy=1'b1", "Configured delay complete"],
        ["WAIT", "default", "WAIT", "busy=1'b1", "Continue waiting"],
        ["DONE", "default", "IDLE", "done=1'b1", "Pulse done and return"],
        ["ERROR", "default", "IDLE", "error=1'b1", "Pulse error and return"],
    ]:
        transitions.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F4CCCC")
    thin_border = Border(bottom=Side(style="thin", color="B7B7B7"))

    sheet_setup = {
        config: ([24, 28, 56], "4472C4"),
        inputs: ([24, 12, 56], "70AD47"),
        outputs: ([24, 12, 24, 56], "ED7D31"),
        states: ([24, 18, 28, 56], "A64D79"),
        transitions: ([24, 38, 24, 62, 56], "5B9BD5"),
    }
    header_help = {
        config: [
            "Configuration key. Keep the provided key names unchanged.",
            "Value used by the generator. Yellow cells are editable.",
            "Human-readable explanation; ignored by the generator.",
        ],
        inputs: [
            "Valid Verilog input port name, for example start or data_valid.",
            "Positive bit width. Use 1 for a scalar and 8 for [7:0].",
            "Optional note for reviewers; ignored by the generator.",
        ],
        outputs: [
            "Valid Verilog output port name, for example busy or done.",
            "Positive bit width. Use 1 for a scalar and 8 for [7:0].",
            "Combinational default, for example 1'b0 or 8'h00.",
            "Optional note for reviewers; ignored by the generator.",
        ],
        states: [
            "Unique Verilog state name. Duplicate names are highlighted red.",
            "Optional explicit encoding, for example 3'b010. Leave blank for automatic encoding.",
            "Optional number/expression of cycles in this state, for example 16 or delay_cycles.",
            "Optional state purpose; ignored by the generator.",
        ],
        transitions: [
            "Current state. Select a name defined in the States sheet.",
            "Verilog condition or default. Rows are evaluated top-to-bottom; put default last.",
            "Destination state. Select a name defined in the States sheet.",
            "Optional branch assignments separated by semicolons, for example busy=1'b1;done=1'b0.",
            "Optional transition purpose; ignored by the generator.",
        ],
    }
    for ws, (widths, tab_color) in sheet_setup.items():
        ws.freeze_panes = "A2"
        editable_rows = 500 if ws is transitions else 200
        last_column = get_column_letter(len(widths))
        ws.auto_filter.ref = f"A1:{last_column}{editable_rows}"
        ws.sheet_properties.tabColor = tab_color
        for index, cell in enumerate(ws[1]):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.comment = Comment(header_help[ws][index], "FSM Generator")
        for row in ws.iter_rows(min_row=2, max_row=editable_rows, max_col=len(widths)):
            for cell in row:
                cell.fill = input_fill
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width

    guide.sheet_properties.tabColor = "00B0F0"
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 110
    guide["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    guide["B1"].font = Font(bold=True, size=12, color="FFFFFF")
    guide["A1"].fill = header_fill
    guide["B1"].fill = header_fill
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border

    bool_validation = DataValidation(type="list", formula1='"true,false"', allow_blank=False)
    config.add_data_validation(bool_validation)
    bool_validation.add(config["B5"])

    width_validation = DataValidation(type="list", formula1='"1,2,4,8,16,32,64"', allow_blank=False)
    width_validation.errorStyle = "warning"
    width_validation.error = "Choose a common width or type another positive width."
    for ws, column in ((inputs, "B"), (outputs, "B")):
        ws.add_data_validation(width_validation)
        width_validation.add(f"{column}2:{column}200")

    state_names = DefinedName("StateNames", attr_text="'States'!$A$2:$A$200")
    wb.defined_names.add(state_names)
    state_validation = DataValidation(type="list", formula1="=StateNames", allow_blank=False)
    transitions.add_data_validation(state_validation)
    state_validation.add("A2:A500")
    next_state_validation = DataValidation(type="list", formula1="=StateNames", allow_blank=False)
    transitions.add_data_validation(next_state_validation)
    next_state_validation.add("C2:C500")
    reset_state_validation = DataValidation(type="list", formula1="=StateNames", allow_blank=False)
    config.add_data_validation(reset_state_validation)
    reset_state_validation.add(config["B6"])

    duplicate_state_rule = FormulaRule(formula=["COUNTIF($A$2:$A$200,A2)>1"], fill=error_fill)
    states.conditional_formatting.add("A2:A200", duplicate_state_rule)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def load_sheets(path: Path, template_path: Optional[Path], sheet_name: Optional[str]) -> Tuple[Dict[str, List[Dict[str, str]]], str]:
    if path.is_dir():
        sheets: Dict[str, List[Dict[str, str]]] = {}
        template = None
        for csv_path in sorted(path.glob("*.csv")):
            key = csv_path.stem.strip().lower()
            if key == "template":
                template = read_template_csv(csv_path)
            else:
                sheets[key] = read_csv_rows(csv_path)
        if sheet_name:
            key = sheet_name.strip().lower()
            if key not in sheets:
                die(f"Folder input has no CSV sheet named '{sheet_name}'")
            sheets = {"table": sheets[key]}
        if template_path:
            template = template_path.read_text(encoding="utf-8")
        return sheets, template or DEFAULT_TEMPLATE

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        sheets, template = read_xlsx_sheets(path)
        if sheet_name:
            key = sheet_name.strip().lower()
            if key not in sheets:
                die(f"Workbook has no data sheet named '{sheet_name}'")
            sheets = {"table": sheets[key]}
            if template_path:
                template = template_path.read_text(encoding="utf-8")
            return sheets, template or DEFAULT_TEMPLATE
        known = {
            "meta",
            "config",
            "states",
            "inputs",
            "outputs",
            "transitions",
            "template",
            "guide",
            "lists",
        }
        data_sheets = [name for name, rows in sheets.items() if name not in known and rows]
        if "transitions" not in sheets and len(data_sheets) == 1:
            sheets = {"table": sheets[data_sheets[0]]}
        elif "transitions" not in sheets and "table" not in sheets and len(data_sheets) > 1:
            die(f"Workbook has multiple data sheets {data_sheets}. Keep one data sheet or use a sheet named transitions.")
        if template_path:
            template = template_path.read_text(encoding="utf-8")
        return sheets, template or DEFAULT_TEMPLATE

    if suffix == ".csv":
        rows = read_csv_rows(path)
        if rows and "sheet" in rows[0]:
            sheets: Dict[str, List[Dict[str, str]]] = {}
            for row in rows:
                sheet = row.pop("sheet", "").strip().lower()
                if sheet:
                    sheets.setdefault(sheet, []).append(row)
            if sheet_name:
                key = sheet_name.strip().lower()
                if key not in sheets:
                    die(f"CSV input has no sheet column value named '{sheet_name}'")
                sheets = {"table": sheets[key]}
        else:
            sheets = {"table": rows}
        template = template_path.read_text(encoding="utf-8") if template_path else DEFAULT_TEMPLATE
        return sheets, template

    die(f"Unsupported input type: {path}")


def pick(row: Dict[str, str], *keys: str, default: str = "") -> str:
    for key in keys:
        value = row.get(key.lower(), "")
        if value != "":
            return value
    return default


def sanitize_ident(name: str, prefix: str = "S") -> str:
    candidate = re.sub(r"[^A-Za-z0-9_$]", "_", name.strip())
    if not candidate:
        candidate = prefix
    if not re.match(r"^[A-Za-z_]", candidate):
        candidate = f"{prefix}_{candidate}"
    return candidate


def require_ident(name: str, context: str) -> str:
    if not IDENT_RE.match(name):
        die(f"Invalid Verilog identifier '{name}' in {context}")
    return name


def parse_meta(rows: List[Dict[str, str]], fallback_module: str) -> Dict[str, str]:
    meta = {
        "module_name": fallback_module,
        "clock": "clk",
        "reset": "rst_n",
        "reset_active_low": "true",
        "reset_state": "",
        "state_reg_name": "state",
        "next_state_name": "next_state",
        "encoding": "binary",
        "output_file": "",
        "timer_width": "32",
        "timer_counter_name": "state_timer",
        "timer_done_name": "timer_done",
        "timer_active_name": "timer_active",
        "language": "systemverilog",
        "reset_type": "async",
        "clock_enable": "",
        "assertions": "true",
    }
    for row in rows:
        key = pick(row, "key", "name")
        value = pick(row, "value")
        if key:
            meta[key.strip().lower()] = value
    return meta


def parse_fields(rows: List[Dict[str, str]], kind: str) -> List[Field]:
    fields = []
    for idx, row in enumerate(rows, start=2):
        name = pick(row, "name", kind)
        if not name:
            continue
        require_ident(name, f"{kind} row {idx}")
        fields.append(
            Field(
                name=name,
                width=pick(row, "width", "bits", default="1"),
                default=pick(row, "default", "reset", default="0"),
            )
        )
    return fields


def split_items(text: str) -> List[str]:
    return [item.strip() for item in re.split(r"[;\n]", text) if item.strip()]


def parse_input_list(text: str) -> List[Field]:
    fields = []
    for item in split_items(text):
        if ":" in item:
            name, width = item.split(":", 1)
        else:
            name, width = item, "1"
        name = name.strip()
        width = width.strip() or "1"
        require_ident(name, "single-table inputs")
        fields.append(Field(name=name, width=width, default="0"))
    return fields


def parse_output_list(text: str) -> List[Field]:
    fields = []
    for item in split_items(text):
        left, default = (item.split("=", 1) + ["0"])[:2] if "=" in item else (item, "0")
        if ":" in left:
            name, width = left.split(":", 1)
        else:
            name, width = left, "1"
        name = name.strip()
        width = width.strip() or "1"
        default = default.strip() or "0"
        require_ident(name, "single-table output_defaults")
        fields.append(Field(name=name, width=width, default=default))
    return fields


def first_non_empty(rows: List[Dict[str, str]], *keys: str) -> str:
    for row in rows:
        value = pick(row, *keys)
        if value:
            return value
    return ""


def build_single_table_spec(rows: List[Dict[str, str]], fallback_module: str, template: str) -> Spec:
    meta = parse_meta([], fallback_module)
    for key in (
        "module_name",
        "clock",
        "reset",
        "reset_active_low",
        "reset_state",
        "state_reg_name",
        "next_state_name",
        "output_file",
        "timer_width",
        "timer_counter_name",
        "timer_done_name",
        "timer_active_name",
        "language",
        "reset_type",
        "clock_enable",
        "assertions",
    ):
        value = first_non_empty(rows, key)
        if value:
            meta[key] = value

    transitions = parse_transitions(rows)

    state_rows_by_name: Dict[str, Dict[str, str]] = {}
    for row in rows:
        state = pick(row, "state", "current_state", "current")
        code = pick(row, "state_code", "code", "encoding")
        timer_limit = pick(row, "timer_limit", "duration", "cycles")
        if not state:
            continue
        if state not in state_rows_by_name:
            state_rows_by_name[state] = {
                "state": state,
                "code": code,
                "timer_limit": timer_limit,
            }
            continue
        state_row = state_rows_by_name[state]
        if code:
            if state_row["code"] and state_row["code"] != code:
                die(f"State '{state}' has conflicting state_code values")
            state_row["code"] = code
        if timer_limit:
            if state_row["timer_limit"] and state_row["timer_limit"] != timer_limit:
                die(f"State '{state}' has conflicting timer_limit values")
            state_row["timer_limit"] = timer_limit
    state_rows = list(state_rows_by_name.values())
    states = parse_states(state_rows, transitions)

    input_text = first_non_empty(rows, "inputs", "input_list")
    output_text = first_non_empty(rows, "output_defaults", "outputs_default", "output_list")
    inputs = parse_input_list(input_text)
    outputs = parse_output_list(output_text)

    if not outputs:
        inferred = []
        seen_outputs = set()
        for tr in transitions:
            for name in tr.outputs:
                if name not in seen_outputs:
                    seen_outputs.add(name)
                    inferred.append(Field(name=name, width="1", default="0"))
        outputs = inferred

    return Spec(meta=meta, states=states, inputs=inputs, outputs=outputs, transitions=transitions, template=template)


def parse_output_assignments(text: str, row_num: int) -> Dict[str, str]:
    result = {}
    if not text.strip():
        return result
    parts = split_items(text)
    for part in parts:
        item = part.strip()
        if not item:
            continue
        if "=" not in item:
            die(f"Transition row {row_num}: output assignment '{item}' must use name=value")
        name, value = item.split("=", 1)
        name = name.strip()
        value = value.strip()
        require_ident(name, f"transition row {row_num} output assignment")
        if not value:
            die(f"Transition row {row_num}: output '{name}' has empty value")
        result[name] = value
    return result


def parse_transitions(rows: List[Dict[str, str]]) -> List[Transition]:
    transitions = []
    for idx, row in enumerate(rows, start=2):
        state = pick(row, "state", "current_state", "current")
        next_state = pick(row, "next_state", "next")
        condition = pick(row, "condition", "cond", "when", default="default")
        outputs = pick(row, "outputs", "assign", "assignments", default="")
        if not state and not next_state:
            continue
        if not state or not next_state:
            die(f"Transition row {idx}: both state and next_state are required")
        transitions.append(
            Transition(
                state=state,
                condition=condition,
                next_state=next_state,
                outputs=parse_output_assignments(outputs, idx),
                row_num=idx,
            )
        )
    return transitions


def parse_states(rows: List[Dict[str, str]], transitions: List[Transition]) -> List[State]:
    raw_states = []
    codes: Dict[str, str] = {}
    timer_limits: Dict[str, str] = {}

    for idx, row in enumerate(rows, start=2):
        name = pick(row, "state", "name")
        if not name:
            continue
        if name not in raw_states:
            raw_states.append(name)
        code = pick(row, "code", "encoding")
        if code:
            if name in codes and codes[name] != code:
                die(f"State '{name}' has conflicting state encoding values")
            codes[name] = code
        timer_limit = pick(row, "timer_limit", "duration", "cycles")
        if timer_limit:
            if name in timer_limits and timer_limits[name] != timer_limit:
                die(f"State '{name}' has conflicting timer_limit values")
            timer_limits[name] = timer_limit

    if not raw_states:
        for tr in transitions:
            if tr.state not in raw_states:
                raw_states.append(tr.state)
            if tr.next_state not in raw_states:
                raw_states.append(tr.next_state)

    states = []
    used_idents = set()
    for raw in raw_states:
        ident = sanitize_ident(raw)
        if not IDENT_RE.match(ident):
            die(f"Cannot sanitize state name '{raw}' into a valid Verilog identifier")
        if ident in used_idents:
            die(f"Duplicate state identifier after sanitize: {ident}")
        used_idents.add(ident)
        states.append(
            State(
                raw_name=raw,
                ident=ident,
                code=codes.get(raw),
                timer_limit=timer_limits.get(raw),
            )
        )

    return states


def build_spec(path: Path, template_path: Optional[Path], sheet_name: Optional[str] = None) -> Spec:
    sheets, template = load_sheets(path, template_path, sheet_name)
    fallback_module = sanitize_ident(path.stem if path.is_file() else path.name, "fsm")
    if "table" in sheets:
        return build_single_table_spec(sheets["table"], fallback_module, template)
    transitions = parse_transitions(sheets.get("transitions", []))
    states = parse_states(sheets.get("states", []), transitions)
    meta = parse_meta(sheets.get("meta", sheets.get("config", [])), fallback_module)
    inputs = parse_fields(sheets.get("inputs", []), "input")
    outputs = parse_fields(sheets.get("outputs", []), "output")
    return Spec(meta=meta, states=states, inputs=inputs, outputs=outputs, transitions=transitions, template=template)


def state_lookup(spec: Spec) -> Dict[str, State]:
    lookup = {}
    for state in spec.states:
        lookup[state.raw_name] = state
        lookup[state.ident] = state
    return lookup


def state_width(spec: Spec) -> int:
    coded_width = 0
    for state in spec.states:
        if state.code:
            match = re.match(r"(\d+)'[bhdBHD]", state.code.strip())
            if match:
                coded_width = max(coded_width, int(match.group(1)))
    if coded_width:
        return coded_width
    encoding = spec.meta.get("encoding", "binary").strip().lower()
    if encoding == "onehot":
        return max(1, len(spec.states))
    return max(1, math.ceil(math.log2(max(1, len(spec.states)))))


def assign_state_codes(spec: Spec) -> None:
    width = state_width(spec)
    encoding = spec.meta.get("encoding", "binary").strip().lower()
    for idx, state in enumerate(spec.states):
        if not state.code:
            if encoding == "onehot":
                state.code = f"{width}'b{1 << idx:0{width}b}"
            elif encoding == "gray":
                state.code = f"{width}'d{idx ^ (idx >> 1)}"
            else:
                state.code = f"{width}'d{idx}"


def timed_states(spec: Spec) -> List[State]:
    return [state for state in spec.states if state.timer_limit]


def parse_bool(value: str, default: bool = False) -> bool:
    text = str(value).strip().lower()
    if not text:
        return default
    if text in ("1", "true", "yes", "on"):
        return True
    if text in ("0", "false", "no", "off"):
        return False
    die(f"Expected a boolean value, got '{value}'")


def is_systemverilog(spec: Spec) -> bool:
    return spec.meta.get("language", "systemverilog") == "systemverilog"


def validate_spec(spec: Spec) -> None:
    if not spec.transitions:
        die("No transitions found. Add transitions.csv or a transitions sheet.")
    if not spec.states:
        die("No states found. Add states.csv or state rows in transitions.csv.")

    require_ident(spec.meta["module_name"], "meta.module_name")
    require_ident(spec.meta["clock"], "meta.clock")
    require_ident(spec.meta["reset"], "meta.reset")
    require_ident(spec.meta["state_reg_name"], "meta.state_reg_name")
    require_ident(spec.meta["next_state_name"], "meta.next_state_name")

    language = spec.meta.get("language", "systemverilog").strip().lower()
    language_aliases = {
        "sv": "systemverilog",
        "systemverilog": "systemverilog",
        "v": "verilog",
        "verilog": "verilog",
        "verilog-2001": "verilog",
    }
    if language not in language_aliases:
        die("language must be systemverilog or verilog")
    spec.meta["language"] = language_aliases[language]

    encoding = spec.meta.get("encoding", "binary").strip().lower()
    if encoding not in ("binary", "onehot", "gray"):
        die("encoding must be binary, onehot, or gray")
    spec.meta["encoding"] = encoding

    reset_type = spec.meta.get("reset_type", "async").strip().lower()
    if reset_type not in ("async", "sync"):
        die("reset_type must be async or sync")
    spec.meta["reset_type"] = reset_type
    parse_bool(spec.meta.get("reset_active_low", "true"), default=True)
    parse_bool(spec.meta.get("assertions", "true"), default=True)

    clock_enable = spec.meta.get("clock_enable", "").strip()
    if clock_enable:
        require_ident(clock_enable, "meta.clock_enable")
        spec.meta["clock_enable"] = clock_enable

    port_names = [spec.meta["clock"], spec.meta["reset"]]
    port_names.extend(item.name for item in spec.inputs)
    port_names.extend(item.name for item in spec.outputs)
    duplicates = sorted({name for name in port_names if port_names.count(name) > 1})
    if duplicates:
        die(f"Duplicate port name: {', '.join(duplicates)}")
    clock_enable_conflicts = {
        spec.meta["clock"],
        spec.meta["reset"],
        *(item.name for item in spec.outputs),
    }
    if clock_enable and clock_enable in clock_enable_conflicts:
        die(f"clock_enable conflicts with existing port '{clock_enable}'")

    try:
        timer_width = int(spec.meta.get("timer_width", "32"), 10)
    except ValueError:
        die("timer_width must be a positive decimal integer")
    if timer_width < 1:
        die("timer_width must be a positive decimal integer")

    for state in timed_states(spec):
        limit = str(state.timer_limit).strip()
        if re.fullmatch(r"-\d+", limit):
            die(f"State '{state.raw_name}' has a negative timer_limit")
        if limit.isdigit() and int(limit, 10) > (1 << timer_width):
            die(
                f"State '{state.raw_name}' timer_limit {limit} does not fit "
                f"timer_width {timer_width}"
            )

    if timed_states(spec):
        timer_names = [
            spec.meta["timer_counter_name"],
            spec.meta["timer_done_name"],
            spec.meta["timer_active_name"],
            "TIMER_W",
        ]
        for name in timer_names:
            require_ident(name, "timer metadata")
        if len(set(timer_names)) != len(timer_names):
            die("Timer internal names must be unique and cannot use the reserved name TIMER_W")
        used_names = {
            spec.meta["clock"],
            spec.meta["reset"],
            spec.meta["state_reg_name"],
            spec.meta["next_state_name"],
            *(item.name for item in spec.inputs),
            *(item.name for item in spec.outputs),
            *(state.ident for state in spec.states),
        }
        conflicts = sorted(set(timer_names) & used_names)
        if conflicts:
            die(f"Timer internal name conflicts with an existing signal: {', '.join(conflicts)}")

        required_placeholders = ("{{timer_declarations}}", "{{timer_block}}")
        missing = [item for item in required_placeholders if item not in spec.template]
        if missing:
            die(
                "Timed states require timer placeholders in the Verilog template: "
                + ", ".join(missing)
            )

    lookup = state_lookup(spec)
    reset_state = spec.meta.get("reset_state", "")
    if not reset_state:
        spec.meta["reset_state"] = spec.states[0].raw_name
    elif reset_state not in lookup:
        die(f"reset_state '{reset_state}' is not listed in states")

    output_names = {item.name for item in spec.outputs}
    for tr in spec.transitions:
        if tr.state not in lookup:
            die(f"Transition row {tr.row_num}: state '{tr.state}' is not listed in states")
        if tr.next_state not in lookup:
            die(f"Transition row {tr.row_num}: next_state '{tr.next_state}' is not listed in states")
        for name in tr.outputs:
            if name not in output_names:
                die(f"Transition row {tr.row_num}: output '{name}' is not listed in outputs")

    seen_default = set()
    seen_condition = set()
    for tr in spec.transitions:
        key = lookup[tr.state].ident
        if tr.is_default:
            if key in seen_default:
                die(f"State '{tr.state}' has more than one default transition")
            seen_default.add(key)
        else:
            cond_key = (key, tr.condition)
            if cond_key in seen_condition:
                die(f"State '{tr.state}' has duplicate condition '{tr.condition}'")
            seen_condition.add(cond_key)

    assign_state_codes(spec)
    codes = [state.code.strip().lower() for state in spec.states if state.code]
    duplicate_codes = sorted({code for code in codes if codes.count(code) > 1})
    if duplicate_codes:
        die(f"Duplicate state encoding: {', '.join(duplicate_codes)}")


def indent(text: str, spaces: int) -> str:
    pad = " " * spaces
    return "\n".join(pad + line if line else line for line in text.splitlines())


def render_ports(spec: Spec) -> str:
    net_type = "logic" if is_systemverilog(spec) else "wire"
    output_type = "logic" if is_systemverilog(spec) else "reg"
    ports = []
    ports.append(f"  input {net_type} {spec.meta['clock']}")
    ports.append(f"  input {net_type} {spec.meta['reset']}")
    clock_enable = spec.meta.get("clock_enable", "")
    input_names = {item.name for item in spec.inputs}
    if clock_enable and clock_enable not in input_names:
        ports.append(f"  input {net_type} {clock_enable}")
    for item in spec.inputs:
        ports.append(f"  input {net_type}{item.decl_width} {item.name}")
    for item in spec.outputs:
        ports.append(f"  output {output_type}{item.decl_width} {item.name}")
    return ",\n".join(ports)


def render_state_localparams(spec: Spec) -> str:
    width = state_width(spec)
    lines = [f"localparam integer STATE_W = {width};"]
    for state in spec.states:
        lines.append(f"localparam [STATE_W-1:0] {state.ident} = {state.code};")
    return "\n".join(lines)


def render_state_declarations(spec: Spec) -> str:
    if not is_systemverilog(spec):
        return "\n".join(
            [
                render_state_localparams(spec),
                "",
                f"reg [STATE_W-1:0] {spec.meta['state_reg_name']};",
                f"reg [STATE_W-1:0] {spec.meta['next_state_name']};",
            ]
        )
    width = state_width(spec)
    members = ",\n".join(f"  {state.ident} = {state.code}" for state in spec.states)
    return "\n".join(
        [
            f"localparam int unsigned STATE_W = {width};",
            "typedef enum logic [STATE_W-1:0] {",
            members,
            "} state_t;",
            "",
            f"state_t {spec.meta['state_reg_name']};",
            f"state_t {spec.meta['next_state_name']};",
        ]
    )


def active_reset_condition(spec: Spec) -> Tuple[str, str]:
    rst = spec.meta["reset"]
    active_low = spec.meta.get("reset_active_low", "true").strip().lower() in ("1", "true", "yes", "low", "active_low")
    if active_low:
        return f"negedge {rst}", f"!{rst}"
    return f"posedge {rst}", rst


def render_state_register_block(spec: Spec, lookup: Dict[str, State]) -> str:
    clk = spec.meta["clock"]
    state_reg = spec.meta["state_reg_name"]
    next_state = spec.meta["next_state_name"]
    reset_state = lookup[spec.meta["reset_state"]].ident
    reset_edge, reset_cond = active_reset_condition(spec)
    always_keyword = "always_ff" if is_systemverilog(spec) else "always"
    sensitivity = f"posedge {clk}"
    if spec.meta.get("reset_type", "async") == "async":
        sensitivity += f" or {reset_edge}"
    clock_enable = spec.meta.get("clock_enable", "")
    lines = [f"{always_keyword} @({sensitivity}) begin", f"  if ({reset_cond}) begin"]
    lines.extend([f"    {state_reg} <= {reset_state};", "  end"])
    if clock_enable:
        lines.extend([f"  else if ({clock_enable}) begin", f"    {state_reg} <= {next_state};", "  end"])
    else:
        lines.extend(["  else begin", f"    {state_reg} <= {next_state};", "  end"])
    lines.append("end")
    return "\n".join(lines)


def render_timer_declarations(spec: Spec) -> str:
    if not timed_states(spec):
        return ""
    width = int(spec.meta["timer_width"], 10)
    counter = spec.meta["timer_counter_name"]
    done = spec.meta["timer_done_name"]
    active = spec.meta["timer_active_name"]
    reg_type = "logic" if is_systemverilog(spec) else "reg"
    param_type = "int unsigned" if is_systemverilog(spec) else "integer"
    return "\n".join(
        [
            f"localparam {param_type} TIMER_W = {width};",
            f"{reg_type} [TIMER_W-1:0] {counter};",
            f"{reg_type} {done};",
            f"{reg_type} {active};",
        ]
    )


def render_timer_block(spec: Spec) -> str:
    states = timed_states(spec)
    if not states:
        return ""

    clk = spec.meta["clock"]
    state_reg = spec.meta["state_reg_name"]
    next_state = spec.meta["next_state_name"]
    counter = spec.meta["timer_counter_name"]
    done = spec.meta["timer_done_name"]
    active = spec.meta["timer_active_name"]
    reset_edge, reset_cond = active_reset_condition(spec)
    seq_keyword = "always_ff" if is_systemverilog(spec) else "always"
    comb_keyword = "always_comb" if is_systemverilog(spec) else "always @(*)"
    sensitivity = f"posedge {clk}"
    if spec.meta.get("reset_type", "async") == "async":
        sensitivity += f" or {reset_edge}"
    clock_enable = spec.meta.get("clock_enable", "")

    lines = [f"{seq_keyword} @({sensitivity}) begin", f"  if ({reset_cond}) begin"]
    lines.extend([f"    {counter} <= {{TIMER_W{{1'b0}}}};", "  end else begin"])
    branch_indent = "    "
    if clock_enable:
        lines.append(f"    if ({clock_enable}) begin")
        branch_indent = "      "
    lines.extend(
        [
            f"{branch_indent}if ({state_reg} != {next_state}) begin",
            f"{branch_indent}  {counter} <= {{TIMER_W{{1'b0}}}};",
            f"{branch_indent}end else if ({active} && !{done}) begin",
            f"{branch_indent}  {counter} <= {counter} + 1'b1;",
            f"{branch_indent}end else if (!{active}) begin",
            f"{branch_indent}  {counter} <= {{TIMER_W{{1'b0}}}};",
            f"{branch_indent}end",
        ]
    )
    if clock_enable:
        lines.append("    end")
    lines.extend([
        "  end",
        "end",
        "",
        f"{comb_keyword} begin",
        f"  {active} = 1'b0;",
        f"  {done} = 1'b0;",
        "",
        f"  {'unique ' if is_systemverilog(spec) else ''}case ({state_reg})",
    ])
    for state in states:
        limit = state.timer_limit
        lines.extend(
            [
                f"    {state.ident}: begin",
                f"      {active} = 1'b1;",
                f"      if (({limit}) <= 1)",
                f"        {done} = 1'b1;",
                "      else",
                f"        {done} = ({counter} >= (({limit}) - 1'b1));",
                "    end",
            ]
        )
    lines.extend(["    default: begin", "    end", "  endcase", "end"])
    return "\n".join(lines)


def output_default_lines(spec: Spec, spaces: int = 2) -> List[str]:
    pad = " " * spaces
    return [f"{pad}{item.name} = {item.default};" for item in spec.outputs]


def assignment_lines(spec: Spec, tr: Transition, lookup: Dict[str, State], spaces: int) -> List[str]:
    pad = " " * spaces
    lines = [f"{pad}{spec.meta['next_state_name']} = {lookup[tr.next_state].ident};"]
    for name, value in tr.outputs.items():
        lines.append(f"{pad}{name} = {value};")
    return lines


def render_state_branch(spec: Spec, state: State, transitions: List[Transition], lookup: Dict[str, State]) -> str:
    state_transitions = [tr for tr in transitions if lookup[tr.state].ident == state.ident]
    conditional = [tr for tr in state_transitions if not tr.is_default]
    default_tr = next((tr for tr in state_transitions if tr.is_default), None)
    next_name = spec.meta["next_state_name"]

    lines = [f"    {state.ident}: begin"]
    if conditional:
        for idx, tr in enumerate(conditional):
            keyword = "if" if idx == 0 else "else if"
            lines.append(f"      {keyword} ({tr.condition}) begin")
            lines.extend(assignment_lines(spec, tr, lookup, 8))
            lines.append("      end")
        lines.append("      else begin")
        if default_tr:
            lines.extend(assignment_lines(spec, default_tr, lookup, 8))
        else:
            lines.append(f"        {next_name} = {state.ident};")
        lines.append("      end")
    elif default_tr:
        lines.extend(assignment_lines(spec, default_tr, lookup, 6))
    else:
        lines.append(f"      {next_name} = {state.ident};")
    lines.append("    end")
    return "\n".join(lines)


def render_comb_block(spec: Spec, lookup: Dict[str, State]) -> str:
    state_reg = spec.meta["state_reg_name"]
    next_state = spec.meta["next_state_name"]
    reset_state = lookup[spec.meta["reset_state"]].ident
    comb_keyword = "always_comb" if is_systemverilog(spec) else "always @(*)"
    lines = [f"{comb_keyword} begin"]
    lines.append(f"  {next_state} = {state_reg};")
    lines.extend(output_default_lines(spec, 2))
    lines.append("")
    lines.append(f"  {'unique ' if is_systemverilog(spec) else ''}case ({state_reg})")
    for state in spec.states:
        lines.append(render_state_branch(spec, state, spec.transitions, lookup))
    lines.append("    default: begin")
    lines.append(f"      {next_state} = {reset_state};")
    lines.append("    end")
    lines.append("  endcase")
    lines.append("end")
    return "\n".join(lines)


def render_assertion_block(spec: Spec) -> str:
    if not is_systemverilog(spec) or not parse_bool(spec.meta.get("assertions", "true"), default=True):
        return ""
    clk = spec.meta["clock"]
    state_reg = spec.meta["state_reg_name"]
    _, reset_cond = active_reset_condition(spec)
    legal_states = ", ".join(state.ident for state in spec.states)
    module_name = spec.meta["module_name"]
    return "\n".join(
        [
            "`ifndef SYNTHESIS",
            "// Simulation-only checks for X propagation and corrupted state encodings.",
            f"assert property (@(posedge {clk}) disable iff ({reset_cond}) !$isunknown({state_reg}))",
            f'  else $error("{module_name}: state contains X/Z");',
            f"assert property (@(posedge {clk}) disable iff ({reset_cond}) {state_reg} inside {{{legal_states}}})",
            f'  else $error("{module_name}: illegal state encoding");',
            "`endif",
        ]
    )


def render(spec: Spec) -> str:
    validate_spec(spec)
    lookup = state_lookup(spec)
    template = spec.template
    if template == DEFAULT_SYSTEMVERILOG_TEMPLATE and not is_systemverilog(spec):
        template = DEFAULT_VERILOG_TEMPLATE
    replacements = {
        "module_name": spec.meta["module_name"],
        "port_declarations": render_ports(spec),
        "state_localparams": render_state_localparams(spec),
        "state_declarations": render_state_declarations(spec),
        "state_reg_name": spec.meta["state_reg_name"],
        "next_state_name": spec.meta["next_state_name"],
        "timer_declarations": render_timer_declarations(spec),
        "state_register_block": render_state_register_block(spec, lookup),
        "timer_block": render_timer_block(spec),
        "comb_block": render_comb_block(spec, lookup),
        "assertion_block": render_assertion_block(spec),
    }
    text = template
    for key, value in replacements.items():
        text = text.replace("{{" + key + "}}", value)
    return text.rstrip() + "\n"


def write_output(spec: Spec, input_path: Path, out_arg: Optional[Path], text: str) -> Path:
    extension = ".sv" if is_systemverilog(spec) else ".v"
    if out_arg:
        if out_arg.is_dir() or str(out_arg).endswith("/") or out_arg.suffix == "":
            out_path = out_arg / f"{spec.meta['module_name']}{extension}"
        else:
            out_path = out_arg
    elif spec.meta.get("output_file"):
        out_path = input_path.parent / spec.meta["output_file"]
    else:
        out_path = Path.cwd() / f"{spec.meta['module_name']}{extension}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(text, encoding="utf-8")
    return out_path


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Generate a synthesizable SystemVerilog or Verilog FSM from CSV/XLSX."
    )
    parser.add_argument("input", nargs="?", type=Path, help="Spec folder, .csv file, or .xlsx workbook")
    parser.add_argument("-o", "--out", type=Path, help="Output HDL file or output directory")
    parser.add_argument("--sheet", help="Data sheet name for .xlsx, folder CSV, or CSV files with a sheet column")
    parser.add_argument("--template", type=Path, help="Optional template file overriding template sheet")
    parser.add_argument("--print", action="store_true", help="Print generated RTL to stdout")
    parser.add_argument(
        "--language",
        choices=("systemverilog", "verilog"),
        help="Override Config.language (default: systemverilog)",
    )
    parser.add_argument(
        "--new-template",
        "--new-adc-template",
        dest="new_template",
        type=Path,
        help="Create a generic FSM workbook template and exit",
    )
    args = parser.parse_args(argv)

    if args.new_template:
        create_fsm_template(args.new_template)
        print(f"Created {args.new_template}")
        return 0

    if not args.input:
        parser.error("input is required unless --new-template is used")

    spec = build_spec(args.input, args.template, args.sheet)
    if args.language:
        spec.meta["language"] = args.language
    text = render(spec)
    if args.print:
        print(text, end="")
        return 0
    out_path = write_output(spec, args.input, args.out, text)
    print(f"Generated {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
